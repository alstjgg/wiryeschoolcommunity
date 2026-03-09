"""Excel 파일 파싱 — 입금내역(.xls/.xlsx) 및 신청자 목록(HTML .xls)"""

import io
import re
from openpyxl import load_workbook


def parse_bank_statement(file_bytes: bytes) -> list[dict]:
    """입금내역 파일 파싱 → 입금 거래 리스트 반환

    포맷:
    - Row 1~6: 메타정보 (스킵)
    - Row 7: 헤더 (거래일시 | 적요 | 의뢰인/수취인 | 입금 | 출금 | ...)
    - Row 8+: 거래 데이터
    - 마지막 행: 합계 (스킵)

    비고 컬럼 없음. 의뢰인은 index 2, 입금은 index 3.
    """
    # .xls (OLE2) → xlrd, .xlsx → openpyxl
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        # openpyxl 실패 → xlrd로 시도 (.xls 포맷)
        import xlrd
        xls_wb = xlrd.open_workbook(file_contents=file_bytes)
        xls_ws = xls_wb.sheet_by_index(0)
        rows = [xls_ws.row_values(i) for i in range(xls_ws.nrows)]

    if len(rows) < 8:
        return []

    # Row 7 (index 6) = 헤더, Row 8+ = 데이터, 마지막 행 = 합계 제외
    data_rows = rows[7:-1]  # index 7부터 마지막 전까지

    transactions = []
    for row in data_rows:
        if len(row) < 4:
            continue

        거래일시 = str(row[0]).strip() if row[0] else ""
        적요 = str(row[1]).strip() if row[1] else ""
        의뢰인 = str(row[2]).strip() if row[2] else ""

        # 입금액 파싱 (index 3)
        raw_amount = row[3]
        if raw_amount is None or str(raw_amount).strip() == "":
            continue
        try:
            입금 = int(float(str(raw_amount).replace(",", "")))
        except (ValueError, TypeError):
            continue

        if 입금 <= 0:
            continue

        transactions.append({
            "거래일시": 거래일시,
            "적요": 적요,
            "의뢰인": 의뢰인,
            "입금": 입금,
        })

    return transactions


def parse_applicant_list(file_bytes: bytes) -> list[dict]:
    """배움숲 신청자 목록(LEARNING_APPLY*.xls, HTML 형식) 파싱

    CRITICAL: HTML 테이블은 23개 헤더 컬럼이지만 데이터 행은 22개 셀.
    '실제결제금액' 컬럼(index 5)이 헤더에만 존재하고 데이터가 없어
    이후 컬럼이 1칸씩 밀림. 보정 필요.

    헤더(23개):
    번호(0), 회차(1), 강좌명(2), 감면정보(3), 수강료(4), 실제결제금액(5),
    신청자(아이디)(6), 성별(7), 생년월일(8), 나이(9), 연락처(10),
    주소(11), 행정동(12), 이메일(13), 분류(14), 교육기간(15),
    신청상태(16), 진행상태(17), 환불신청일(18), 환불은행(19),
    환불계좌번호(20), 환불예금주(21), 환불사유(22)

    데이터(22개, 실제결제금액 누락):
    번호(0), 회차(1), 강좌명(2), 감면정보(3), 수강료(4),
    신청자(아이디)(5→보정후6), 성별(6→7), 생년월일(7→8), 나이(8→9),
    연락처(9→10), 주소(10→11), 행정동(11→12), 이메일(12→13),
    분류(13→14), 교육기간(14→15), 신청상태(15→16), 진행상태(16→17), ...
    """
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(file_bytes, "html.parser")
    table = soup.find("table")
    if not table:
        return []

    all_rows = table.find_all("tr")
    if len(all_rows) < 2:
        return []

    # 첫 행 = 헤더
    header_cells = all_rows[0].find_all(["th", "td"])
    header_count = len(header_cells)

    applicants = []
    for tr in all_rows[1:]:
        cells = [td.get_text(strip=True) for td in tr.find_all(["th", "td"])]
        if not cells or len(cells) < 16:
            continue

        # 실제결제금액(index 5) 데이터 누락 보정: 빈 값 삽입
        if len(cells) < header_count:
            cells.insert(5, "")

        # 인덱스는 보정 후 헤더 기준
        회차 = cells[1]
        강좌명 = cells[2]
        신청자_raw = cells[6]  # "고아라(bnmaam)" 형태
        성별 = cells[7]
        나이 = cells[9]
        연락처 = cells[10]
        주소 = cells[11]
        행정동 = cells[12]
        이메일 = cells[13]
        신청상태 = cells[16]

        # 신청자에서 이름 추출 (괄호 앞 부분)
        name_match = re.match(r"^([^(]+)", 신청자_raw)
        이름 = name_match.group(1).strip() if name_match else 신청자_raw.strip()

        # 전화번호에서 이름ID 생성
        phone_clean = 연락처.replace("-", "").replace(" ", "")
        phone_suffix = phone_clean[-4:] if len(phone_clean) >= 4 else ""
        이름ID = f"{이름}{phone_suffix}" if phone_suffix else 이름

        applicants.append({
            "이름ID": 이름ID,
            "이름": 이름,
            "강좌명": 강좌명,
            "전화번호": 연락처,
            "성별": 성별,
            "나이": 나이,
            "주소": 주소,
            "행정동": 행정동,
            "이메일": 이메일,
            "신청상태": 신청상태,
        })

    return applicants
